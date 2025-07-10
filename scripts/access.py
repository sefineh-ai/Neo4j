import mysql.connector  # type: ignore
import os
from dotenv import load_dotenv  # type: ignore
from openpyxl import Workbook  # For creating Excel files
from openpyxl.utils.exceptions import IllegalCharacterError


load_dotenv()

connection = mysql.connector.connect(
    host=os.getenv("DB_HOST"),
    user=os.getenv("DB_USER"),
    password=os.getenv("DB_PASSWORD"),
    database=os.getenv("DB_NAME"),
    port=int(os.getenv("DB_PORT"))
)

cursor = connection.cursor()

def sanitize_cell_value(value):
    """
    Sanitizes cell values to remove illegal characters for Excel.
    """
    if isinstance(value, bytes):
        value = value.decode('utf-8', errors='replace')
    if isinstance(value, str):
        value = ''.join(c for c in value if ord(c) >= 32)
    return value

try:
    cursor.execute("SHOW TABLES;")
    tables = cursor.fetchall() 

    for (table_name,) in tables: 
        print(f"Exporting table: {table_name}")
        query = f"SELECT * FROM {table_name};"
        cursor.execute(query)
        rows = cursor.fetchall()
        column_names = [i[0] for i in cursor.description]
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = table_name

        for col_num, column_name in enumerate(column_names, start=1):
            sheet.cell(row=1, column=col_num, value=column_name)

        for row_num, row_data in enumerate(rows, start=2):
            for col_num, cell_value in enumerate(row_data, start=1):
                sanitized_value = sanitize_cell_value(cell_value)
                try:
                    sheet.cell(row=row_num, column=col_num, value=sanitized_value)
                except IllegalCharacterError:
                    sheet.cell(row=row_num, column=col_num, value="INVALID_CHAR")

        output_file = f"data/excel/{table_name}.xlsx"
        workbook.save(output_file)
        print(f"Table {table_name} exported successfully to {output_file}")

except mysql.connector.Error as err:
    print(f"Error: {err}")
finally:
    cursor.close()
    connection.close()
